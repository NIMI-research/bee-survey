import csv
import argparse
import logging
import os
import re
import openpyxl

from const import MAIN_CSV_PATH, SOURCE_WORKBOOK_PATH

"""
Standalone script to build Main CSV.

Rules:
- Merge subset sheets into one dataset.
- Subset sheets define output schema (including new columns/rows).
- Compare against existing Sources sheet only for reconciliation.
- On cell mismatch for same row, Sources value is ground truth.
- Do not add columns that exist only in Sources.
"""

logger = logging.getLogger(__name__)


SOURCE_SHEETS = [
    "New-original 2024-2025",
    "Pass",
    "Original",
    "original-2011-2023",
]
SOURCES_SHEET = "Sources"

COLUMN_CANDIDATES = {
    "Paper Title": ["Paper Title", "PaperTitle", "Paper Title of Sources"],
    "Papers": ["Papers", "Paper", "Title"],
    "URL": ["URL", "Link", "Paper URL", "Source URL"],
    "Authors": ["Authors", "Author"],
    "Year": ["Year", "Publication Year"],
    "Venue Name": ["Venue Name", "Venue"],
    "Data modality": ["Data modality", "Data Modality", "Modality", "Data Type"],
    "Pass/Fail": ["Pass/Fail", "Pass Fail"],
    "Search Keyword": ["Search Keyword", "Search Keywords", "Keyword", "Keywords"],
    "Approach (AI)": ["Approach (AI)", "Approach", "AI Approach", "Method", "Methodology"],
    "Abstract": ["Abstract", "Summary"],
    "Research Problem": ["Research Problem", "Problem", "Research Task"],
    "Category": ["Category", "Categories"],
    "Bee Demographic": ["Bee Demographic", "Bee Demographics", "Bee demographic"],
    "Research Demographic": [
        "Research Demographic",
        "Research Demographics",
        "Study Demographic",
        "Study Demographics",
    ],
    "Collection Time": ["Collection Time", "Time of Collection", "Sampling Time"],
    "Approach Group": ["Approach Group", "Approach group", "ApproachGroup"],
}

BASE_COLUMNS = [
    "Paper Title",
    "Papers",
    "URL",
    "Authors",
    "Year",
    "Abstract",
    "Venue Name",
    "Data modality",
    "Pass/Fail",
    "Search Keyword",
    "Approach (AI)",
    "Research Problem",
    "Category",
    "Bee Demographic",
    "Research Demographic",
    "Collection Time",
    "Approach Group",
]

def _normalize_header(name):
    if name is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def _to_int_year(value, default=-1):
    if value is None:
        return default

    text = str(value).strip()
    if not text:
        return default

    year_matches = re.findall(r"(?:19|20)\d{2}", text)
    if year_matches:
        return max(int(y) for y in year_matches)

    any_int = re.search(r"-?\d+", text)
    if any_int:
        try:
            return int(any_int.group(0))
        except ValueError:
            return default

    return default


def _coerce_year_for_output(value):
    parsed = _to_int_year(value, default=None)
    return "" if parsed is None else parsed


def _text(value):
    return "" if value is None else str(value).strip()

def _find_column_index(headers, candidates):
    normalized = {_normalize_header(h): idx + 1 for idx, h in enumerate(headers)}
    for candidate in candidates:
        candidate_key = _normalize_header(candidate)
        if candidate_key in normalized:
            return normalized[candidate_key]
    return None


def _build_dedupe_key(row):
    return (_text(row.get("Papers")).lower(), _to_int_year(row.get("Year")))


def _ensure_columns(columns, row):
    for col in columns:
        row.setdefault(col, "")
    return row


def _build_subset_column_map(sheet):
    headers = [cell.value for cell in sheet[1]]
    index_to_col = {}

    for canonical in BASE_COLUMNS:
        idx = _find_column_index(headers, [canonical, *COLUMN_CANDIDATES.get(canonical, [])])
        if idx:
            index_to_col[idx] = canonical

    return index_to_col


def _build_sources_column_map(sheet, target_columns):
    headers = [cell.value for cell in sheet[1]]
    index_to_col = {}
    for col in target_columns:
        if col in COLUMN_CANDIDATES:
            idx = _find_column_index(headers, [col, *COLUMN_CANDIDATES[col]])
        else:
            idx = _find_column_index(headers, [col])
        if idx:
            index_to_col[idx] = col
    return index_to_col


def _collect_rows(sheet, index_to_col, target_columns):
    rows = []
    for row_idx in range(2, sheet.max_row + 1):
        row = {col: "" for col in target_columns}
        has_value = False
        for idx, col in index_to_col.items():
            value = _text(sheet.cell(row=row_idx, column=idx).value)
            if value:
                has_value = True
            row[col] = value
        if not has_value:
            continue
        if not _text(row.get("Papers")):
            continue
        if "Year" in row:
            row["Year"] = _coerce_year_for_output(row.get("Year"))
        rows.append(row)
    return rows


def _merge_subset_rows(rows, columns):
    merged = {}
    for row in rows:
        row = _ensure_columns(columns, row)
        key = _build_dedupe_key(row)
        if key not in merged:
            merged[key] = row.copy()
            continue
        existing = merged[key]
        for col in columns:
            if not existing.get(col) and row.get(col):
                existing[col] = row[col]
    return merged


def _reconcile_with_sources(merged_rows, source_rows, columns):
    for source_row in source_rows:
        source_row = _ensure_columns(columns, source_row)
        key = _build_dedupe_key(source_row)
        if key not in merged_rows:
            merged_rows[key] = source_row.copy()
            continue
        existing = merged_rows[key]
        for col in columns:
            source_val = _text(source_row.get(col))
            if source_val:
                existing[col] = source_val


def build_main_csv(
    input_path=SOURCE_WORKBOOK_PATH,
    output_path=MAIN_CSV_PATH,
):
    """Build consolidated Main CSV from subset sheets, with Sources reconciliation."""
    workbook = openpyxl.load_workbook(input_path)

    missing_sheets = [name for name in SOURCE_SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"Missing required sheets (case-sensitive): {missing_sheets}")

    columns = list(BASE_COLUMNS)
    merged_rows = []
    for sheet_name in SOURCE_SHEETS:
        sheet = workbook[sheet_name]
        index_to_col = _build_subset_column_map(sheet)
        sheet_rows = _collect_rows(sheet, index_to_col, columns)
        merged_rows.extend(sheet_rows)

    unique_rows = _merge_subset_rows(merged_rows, columns)

    if SOURCES_SHEET in workbook.sheetnames:
        source_sheet = workbook[SOURCES_SHEET]
        source_index_to_col = _build_sources_column_map(source_sheet, columns)
        source_rows = _collect_rows(source_sheet, source_index_to_col, columns)
        _reconcile_with_sources(unique_rows, source_rows, columns)

    final_rows = list(unique_rows.values())
    final_rows.sort(
        key=lambda item: (
            -_to_int_year(item.get("Year")),
            _text(item.get("Papers")).lower(),
        )
    )

    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows([_ensure_columns(columns, row) for row in final_rows])

    return output_path


def main():
    """Run data-builder as a standalone script and save the consolidated CSV."""
    parser = argparse.ArgumentParser(
        description="Build consolidated main CSV from workbook sheets."
    )
    parser.add_argument(
        "--input",
        default=str(SOURCE_WORKBOOK_PATH),
        help="Path to input workbook (.xlsx).",
    )
    parser.add_argument(
        "--output",
        default=str(MAIN_CSV_PATH),
        help="Path to output CSV file.",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    )

    saved_path = build_main_csv(input_path=args.input, output_path=args.output)
    logger.info("Main CSV saved to: %s", saved_path)


if __name__ == "__main__":
    main()
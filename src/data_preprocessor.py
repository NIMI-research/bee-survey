import csv
import argparse
import logging
import re

import openpyxl

from const import MAIN_CSV_PATH, SOURCE_WORKBOOK_PATH

"""
Standalone script to build Main CSV.

Rules (v2):
- Only two sheets are read: "Sources" and "Re-organize".
- No row is discarded. Every row from both sheets ends up in main.csv,
  either merged (if a title match is found across sheets) or standalone.
- Rows are matched by title (fuzzy fallback). Re-organize is ground truth
  for merged columns; its own exclusive columns pass through untouched.
- Sources' exclusive columns pass through untouched.
- If a Re-organize row has no matching Sources row, it is still included,
  and the title + reason is logged so it can be reviewed.
"""

logger = logging.getLogger(__name__)


SOURCES_SHEET = "Sources"
REORGANIZE_SHEET = "Re-organize"

FUZZY_MATCH_THRESHOLD = 0.86

# Raw columns expected in each sheet (normalized against actual headers,
# tolerant of stray whitespace/newlines but not renamed/aliased).
SOURCES_COLUMNS = [
    "Papers",
    "URL",
    "Authors",
    "Year",
    "Abstract",
    "Venue Name",
    "Data modality",
    "Pass/Fail",
    "Search Keyword",
    "Approach (AI)",
    "Research Problem",
    "Category",
    "Bee Demographic",
    "Research Demographic",
    "Collection Time",
    "Approach Group",
]

REORGANIZE_COLUMNS = [
    "Year",
    "Title",
    "Authors",
    "Subcategory (AI Task)",
    "Category (Section)",
    "Approach (AI)",
    "Data Modality",
]

# (sources_col, reorganize_col, output_col) - Re-organize is ground truth
MERGE_PAIRS = [
    ("Year", "Year", "Year"),
    ("Papers", "Title", "Papers"),
    ("Authors", "Authors", "Authors"),
    ("Data modality", "Data Modality", "Data modality"),
    ("Approach (AI)", "Approach (AI)", "Approach (AI)"),
    ("Category", "Subcategory (AI Task)", "Subcategory (AI Task)"),
]
_MERGED_SOURCES_COLS = {p[0] for p in MERGE_PAIRS}
_MERGED_REORG_COLS = {p[1] for p in MERGE_PAIRS}

SOURCES_ONLY_COLUMNS = [c for c in SOURCES_COLUMNS if c not in _MERGED_SOURCES_COLS]
REORGANIZE_ONLY_COLUMNS = [c for c in REORGANIZE_COLUMNS if c not in _MERGED_REORG_COLS]

OUTPUT_COLUMNS = [
    "Papers",
    "URL",
    "Authors",
    "Year",
    "Abstract",
    "Venue Name",
    "Data modality",
    "Pass/Fail",
    "Search Keyword",
    "Approach (AI)",
    "Research Problem",
    "Subcategory (AI Task)",
    "Category (Section)",
    "Bee Demographic",
    "Research Demographic",
    "Collection Time",
    "Approach Group",
]


def _normalize_header(name):
    if name is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def _text(value):
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def _to_int_year(value, default=-1):
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    year_matches = re.findall(r"(?:19|20)\d{2}", text)
    if year_matches:
        return max(int(y) for y in year_matches)
    any_int = re.search(r"-?\d+", text)
    if any_int:
        try:
            return int(any_int.group(0))
        except ValueError:
            return default
    return default


def _coerce_year_for_output(value):
    parsed = _to_int_year(value, default=None)
    return "" if parsed is None else parsed


def _title_key(value):
    text = _text(value).lower()
    text = re.sub(r"\([^)]*\)|\[[^\]]*\]|\{[^}]*\}", " ", text)
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def _fuzzy_similarity(left, right):
    if not left or not right:
        return 0.0
    from difflib import SequenceMatcher
    return SequenceMatcher(None, left, right).ratio()


def _header_index_map(sheet, expected_columns):
    """Map expected column name -> 1-based column index, matched by
    normalized header text (whitespace/newline/case insensitive)."""
    normalized_expected = {_normalize_header(h): h for h in expected_columns}
    headers = [cell.value for cell in sheet[1]]
    result = {}
    for idx, h in enumerate(headers, start=1):
        key = _normalize_header(h)
        if key in normalized_expected:
            result[normalized_expected[key]] = idx
    return result


def _read_sheet_rows(sheet, expected_columns, title_col):
    col_index = _header_index_map(sheet, expected_columns)
    missing = [c for c in expected_columns if c not in col_index]
    raw_headers = [cell.value for cell in sheet[1]]
    logger.info("Sheet '%s' raw header row: %r", sheet.title, raw_headers)
    logger.info("Sheet '%s' resolved column map: %s", sheet.title, col_index)
    if missing:
        logger.warning("Sheet '%s' is missing expected columns: %s", sheet.title, missing)
    if title_col not in col_index:
        logger.error(
            "Sheet '%s' could not locate title column '%s' — every row will be skipped.",
            sheet.title, title_col,
        )

    rows = []
    for r in range(2, sheet.max_row + 1):
        row = {}
        has_value = False
        for col in expected_columns:
            idx = col_index.get(col)
            val = _text(sheet.cell(row=r, column=idx).value) if idx else ""
            if val:
                has_value = True
            row[col] = val
        if not has_value:
            continue
        if not row.get(title_col):
            continue
        rows.append(row)
    return rows


def _match_reorganize_to_sources(reorg_rows, source_rows):
    """Returns (pairs, unmatched_reorg_indices, unmatched_source_indices).
    pairs is a list of (reorg_idx, source_idx)."""
    source_keys = {}
    for sidx, srow in enumerate(source_rows):
        source_keys.setdefault(_title_key(srow["Papers"]), []).append(sidx)

    used_source = set()
    pairs = []
    unmatched_reorg = []

    # exact title-key match first
    for ridx, rrow in enumerate(reorg_rows):
        key = _title_key(rrow["Title"])
        candidates = [s for s in source_keys.get(key, []) if s not in used_source]
        if candidates:
            sidx = candidates[0]
            used_source.add(sidx)
            pairs.append((ridx, sidx))
        else:
            unmatched_reorg.append(ridx)

    # fuzzy fallback for the rest
    still_unmatched = []
    for ridx in unmatched_reorg:
        rkey = _title_key(reorg_rows[ridx]["Title"])
        best_idx, best_score = None, 0.0
        for sidx, srow in enumerate(source_rows):
            if sidx in used_source:
                continue
            score = _fuzzy_similarity(rkey, _title_key(srow["Papers"]))
            if score > best_score:
                best_idx, best_score = sidx, score
        if best_idx is not None and best_score >= FUZZY_MATCH_THRESHOLD:
            used_source.add(best_idx)
            pairs.append((ridx, best_idx))
        else:
            still_unmatched.append(ridx)

    unmatched_source = [s for s in range(len(source_rows)) if s not in used_source]
    return pairs, still_unmatched, unmatched_source


def _merge_row(source_row, reorg_row):
    """source_row/reorg_row may be None. Re-organize wins on merge-pair
    columns when present; each side's exclusive columns pass through."""
    out = {col: "" for col in OUTPUT_COLUMNS}

    for source_col, reorg_col, output_col in MERGE_PAIRS:
        reorg_val = _text(reorg_row.get(reorg_col)) if reorg_row else ""
        source_val = _text(source_row.get(source_col)) if source_row else ""
        value = reorg_val or source_val
        if output_col == "Year":
            value = _coerce_year_for_output(value)
        out[output_col] = value

    for col in SOURCES_ONLY_COLUMNS:
        out[col] = _text(source_row.get(col)) if source_row else ""

    for col in REORGANIZE_ONLY_COLUMNS:
        out[col] = _text(reorg_row.get(col)) if reorg_row else ""

    return out


def build_main_csv(input_path=SOURCE_WORKBOOK_PATH, output_path=MAIN_CSV_PATH):
    """Build main.csv by merging Sources and Re-organize sheets on title.
    No row from either sheet is discarded."""
    workbook = openpyxl.load_workbook(input_path)

    missing_sheets = [s for s in (SOURCES_SHEET, REORGANIZE_SHEET) if s not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"Missing required sheets (case-sensitive): {missing_sheets}")

    source_rows = _read_sheet_rows(workbook[SOURCES_SHEET], SOURCES_COLUMNS, title_col="Papers")
    reorg_rows = _read_sheet_rows(workbook[REORGANIZE_SHEET], REORGANIZE_COLUMNS, title_col="Title")

    pairs, unmatched_reorg, unmatched_source = _match_reorganize_to_sources(reorg_rows, source_rows)

    final_rows = []

    for ridx, sidx in pairs:
        final_rows.append(_merge_row(source_rows[sidx], reorg_rows[ridx]))

    for ridx in unmatched_reorg:
        rrow = reorg_rows[ridx]
        logger.warning(
            "No matching Sources row for Re-organize title: %r (reason: title/fuzzy match below threshold %.2f)",
            rrow.get("Title", ""), FUZZY_MATCH_THRESHOLD,
        )
        final_rows.append(_merge_row(None, rrow))

    for sidx in unmatched_source:
        final_rows.append(_merge_row(source_rows[sidx], None))

    final_rows.sort(
        key=lambda item: (
            -_to_int_year(item.get("Year")),
            item.get("Papers", "").lower(),
        )
    )

    output_dir = output_path.parent if hasattr(output_path, "parent") else None
    if output_dir and not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(final_rows)

    logger.info(
        "Main CSV built: %d merged, %d Re-organize-only, %d Sources-only (total %d rows)",
        len(pairs), len(unmatched_reorg), len(unmatched_source), len(final_rows),
    )

    return output_path


def main():
    parser = argparse.ArgumentParser(description="Build main CSV from Sources + Re-organize sheets.")
    parser.add_argument("--input", default=str(SOURCE_WORKBOOK_PATH), help="Path to input workbook (.xlsx).")
    parser.add_argument("--output", default=str(MAIN_CSV_PATH), help="Path to output CSV file.")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")

    saved_path = build_main_csv(input_path=args.input, output_path=args.output)
    logger.info("Main CSV saved to: %s", saved_path)


if __name__ == "__main__":
    main()